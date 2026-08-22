#!/usr/bin/env python3
"""Export the curated CIS 7000 workbook into browser-friendly JSON and JS."""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


def split_lines(value: object) -> list[str]:
    return [line.strip() for line in str(value or "").splitlines() if line.strip()]


def parse_slot(row: tuple[object, ...]) -> dict[str, object]:
    reading_lines = split_lines(row[6])
    venue_lines = split_lines(row[7])
    source_lines = split_lines(row[14])
    if len(reading_lines) < 2:
        raise ValueError(f"Slot {row[0]} has no paper title")

    focus = reading_lines[0]
    paper_lines = reading_lines[1:]
    papers: list[dict[str, str]] = []
    for index, title_line in enumerate(paper_lines):
        if title_line.startswith("Lead — "):
            role = "Lead"
            title = title_line.removeprefix("Lead — ")
        elif title_line.startswith("Skim — "):
            role = "Skim"
            title = title_line.removeprefix("Skim — ")
        else:
            role = "Lead"
            title = title_line
        papers.append(
            {
                "role": role,
                "title": title,
                "venue": venue_lines[index] if index < len(venue_lines) else "",
                "url": source_lines[index] if index < len(source_lines) else "",
            }
        )

    if not (len(papers) == len(venue_lines) == len(source_lines)):
        raise ValueError(f"Slot {row[0]} paper/venue/source counts do not match")

    return {
        "slotId": row[0],
        "week": int(row[1]),
        "topic": row[2],
        "format": row[3],
        "era": row[4],
        "track": row[5],
        "focus": focus,
        "readingLoad": row[8],
        "discussionHook": row[13],
        "papers": papers,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("output_dir", type=Path)
    args = parser.parse_args()

    workbook = load_workbook(args.workbook, data_only=True, read_only=True)
    signup = workbook["2026 Sign-up"]
    slots = [parse_slot(row) for row in signup.iter_rows(min_row=7, max_row=48, min_col=1, max_col=15, values_only=True)]

    pool_sheet = workbook["Candidate Pool"]
    pool: list[dict[str, object]] = []
    for row in pool_sheet.iter_rows(min_row=5, max_row=128, min_col=1, max_col=13, values_only=True):
        if not row[4]:
            continue
        pool.append(
            {
                "topic": row[0],
                "tier": row[1],
                "era": row[2],
                "tags": split_lines(row[3]) if "\n" in str(row[3] or "") else [tag.strip() for tag in str(row[3] or "").split(";") if tag.strip()],
                "title": row[4],
                "venue": row[5],
                "review": row[6],
                "signal": row[7],
                "contribution": row[8],
                "teachingHook": row[9],
                "rationale": row[10],
                "url": row[11],
                "placement": row[12],
            }
        )

    stats = {
        "slots": len(slots),
        "bundles": sum(len(slot["papers"]) > 1 for slot in slots),
        "classicCentered": sum(slot["era"] == "Classic-centered" for slot in slots),
        "contains2025or2026": sum(slot["era"] != "Classic-centered" for slot in slots),
        "robotics": sum(slot["track"] == "Robotics" for slot in slots),
        "gaussianSplatting": sum(slot["track"] == "GS-centered" for slot in slots),
        "frontier": sum(slot["era"] == "Frontier (2026)" for slot in slots),
        "candidatePapers": len(pool),
        "backupChoices": len(pool) - len(slots),
    }

    if stats != {
        "slots": 42,
        "bundles": 14,
        "classicCentered": 6,
        "contains2025or2026": 36,
        "robotics": 15,
        "gaussianSplatting": 3,
        "frontier": 2,
        "candidatePapers": 124,
        "backupChoices": 82,
    }:
        raise ValueError(f"Unexpected course stats: {stats}")

    payload = {"stats": stats, "slots": slots}
    args.output_dir.mkdir(parents=True, exist_ok=True)
    (args.output_dir / "course-slots.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    (args.output_dir / "paper-pool.json").write_text(json.dumps(pool, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    (args.output_dir / "course-slots.js").write_text("window.COURSE_DATA = " + json.dumps(payload, ensure_ascii=False, indent=2) + ";\n", encoding="utf-8")
    (args.output_dir / "paper-pool.js").write_text("window.PAPER_POOL = " + json.dumps(pool, ensure_ascii=False, indent=2) + ";\n", encoding="utf-8")


if __name__ == "__main__":
    main()
